from openpyxl.workbook import Workbook
import pandas as pd
from datetime import datetime

class WorkbookData:
    EXCEL_FILE_PATH = 'Data/FinanceTrackerData.xlsx'
    COLUMNS = ["Source", "Amount", "Tags", "Date"]
    
    def __init__(self):
        try:
            # Try to read the Excel file
            self.wb = pd.ExcelFile(self.EXCEL_FILE_PATH)
        except FileNotFoundError:
            # Create new Excel file with empty sheets and headers
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create Income sheet with headers
            ws_income = wb.create_sheet('Income')
            ws_income.append(self.COLUMNS)
            
            # Create Expense sheet with headers
            ws_expense = wb.create_sheet('Expense')
            ws_expense.append(self.COLUMNS)
            
            wb.save(self.EXCEL_FILE_PATH)
            self.wb = pd.ExcelFile(self.EXCEL_FILE_PATH)

    def get_sheet_data(self, sheet_name):
        """Get all data from specified sheet using pandas"""
        try:
            # Read the sheet into a DataFrame, don't skip any rows now
            df = pd.read_excel(self.wb, sheet_name=sheet_name)
            
            # Remove empty rows
            df = df.dropna(how='all')
            
            # Skip the header row
            if len(df) > 1:
                df = df.iloc[1:]
            
            # Format amount column
            if 'Amount' in df.columns:
                df['Amount'] = df['Amount'].apply(lambda x: f"₱{float(x):,.2f}" if pd.notnull(x) else x)
            
            # Format date column
            if 'Date' in df.columns:
                df['Date'] = pd.to_datetime(df['Date']).dt.strftime("%B %d, %Y")
            
            # Convert DataFrame to list of tuples
            return df.values.tolist()
            
        except Exception as e:
            print(f"Error reading sheet {sheet_name}: {e}")
            return []

    def save_workbook(self):
        """Save the workbook"""
        # If you need to save changes, implement using pandas to_excel
        pass

    